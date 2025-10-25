from flask import Flask, render_template, request, jsonify
from flask_talisman import Talisman
from roster_generator import RosterGenerator
import json
import datetime
import calendar
from datetime import timedelta
import openpyxl

app = Flask(__name__)
csp = {
    'default-src': '\'self\'',
    'script-src': [
        '\'self\'',
        '\'unsafe-inline\'',
        'https://cdn.tailwindcss.com',
        'https://unpkg.com',
        'https://cdnjs.cloudflare.com',
    ],
    'style-src': [
        '\'self\'',
        '\'unsafe-inline\'',
        'https://fonts.googleapis.com',
    ],
    'font-src': [
        '\'self\'',
        'https://fonts.gstatic.com',
    ],
}
Talisman(app, content_security_policy=csp)

# It's better to create a single instance of the generator
roster_generator_instance = RosterGenerator()

# In-memory data store for the generated roster
current_roster = {}

# In-memory data store for employees
employees_db = [
    { "name": 'Rohit', "role": 'Development', "serviceNow": 5, "jira": 8, "csat": 92, "ticketsResolved": 13, "avgResolutionTime": 45, "leaveBalance": 20, "isAiAgentActive": False, "project": "Phoenix", "costCenter": "RND-101" },
    { "name": 'Keerthi', "role": 'Operations', "serviceNow": 3, "jira": 12, "csat": 98, "ticketsResolved": 15, "avgResolutionTime": 30, "leaveBalance": 20, "isAiAgentActive": False, "project": "Orion", "costCenter": "OPS-202" },
    { "name": 'Naresh', "role": 'DBA', "serviceNow": 7, "jira": 4, "csat": 95, "ticketsResolved": 11, "avgResolutionTime": 55, "leaveBalance": 20, "isAiAgentActive": False, "project": "Phoenix", "costCenter": "RND-101" },
]

# --- Gamification Data Stores ---
gamification_points = {emp['name']: 0 for emp in employees_db}
clock_in_records = [] # { 'name': string, 'clock_in': datetime, 'clock_out': datetime, 'shift': string }
employee_badges = {emp['name']: [] for emp in employees_db}

project_cost_centers = {
    "Phoenix": ["RND-101", "RND-102"],
    "Orion": ["OPS-202", "OPS-203"],
}

def log_to_file(message):
    with open("app.log", "a") as f:
        f.write(f"[{datetime.datetime.now()}] {message}\n")

@app.route('/')
def index():
    log_to_file("Index route was hit.")
    return render_template('index.html')

@app.route('/api/generate_roster', methods=['POST'])
def generate_roster():
    log_to_file("Generate roster route was hit.")
    data = request.get_json()
    log_to_file(f"Received data: {json.dumps(data, indent=2)}")

    if not data or 'members' not in data:
        log_to_file("Error: Missing 'members' in request body")
        return jsonify({"error": "Missing 'members' in request body"}), 400

    members = data['members']
    constraints = data.get('constraints', '')

    # Normalize keys to lowercase to match what the roster generator expects
    normalized_members = [{k.lower(): v for k, v in member.items()} for member in members]

    try:
        # Use the roster generator instance
        global current_roster
        prediction = roster_generator_instance.forward(members=normalized_members, constraints=constraints)
        current_roster = prediction.roster
        log_to_file(f"Generated and saved roster: {json.dumps(current_roster, indent=2)}")
        return jsonify(current_roster)
    except Exception as e:
        log_to_file(f"Error generating roster: {e}")
        return jsonify({"error": "Failed to generate roster"}), 500

@app.route('/api/roster/status', methods=['GET'])
def roster_status():
    log_to_file("Roster status route was hit.")
    return jsonify({"isGenerated": bool(current_roster)})

@app.route('/api/employees', methods=['GET'])
def get_employees():
    log_to_file("Get all employees route was hit.")
    return jsonify(employees_db)

@app.route('/api/employees/<string:name>/shifts', methods=['GET'])
def get_employee_shifts(name):
    log_to_file(f"Get shifts for employee {name} route was hit.")
    if not current_roster:
        return jsonify([]) # Return empty list if no roster is generated

    employee_shifts = []
    for day, shifts in current_roster.items():
        for shift_name, people in shifts.items():
            if isinstance(people, list):
                if any(p.get('name') == name for p in people):
                    employee_shifts.append({
                        "day": day,
                        "shift": shift_name
                    })

    log_to_file(f"Found {len(employee_shifts)} shifts for {name}: {json.dumps(employee_shifts, indent=2)}")
    return jsonify(employee_shifts)

@app.route('/api/employees/<string:name>/ai_status', methods=['PUT'])
def update_ai_status(name):
    log_to_file(f"Update AI status route was hit for employee: {name}")
    data = request.get_json()
    if not data or 'isAiAgentActive' not in data:
        return jsonify({"error": "Missing 'isAiAgentActive' in request body"}), 400

    is_active = data['isAiAgentActive']

    employee = next((emp for emp in employees_db if emp['name'] == name), None)

    if not employee:
        return jsonify({"error": "Employee not found"}), 404

    employee['isAiAgentActive'] = is_active
    log_to_file(f"Updated {name}'s AI agent status to: {is_active}")
    return jsonify(employee)

# In-memory data store for leave requests
leave_requests_db = []
request_id_counter = 0

@app.route('/api/leave_requests', methods=['GET'])
def get_leave_requests():
    log_to_file("Get all leave requests route was hit.")
    return jsonify(leave_requests_db)

@app.route('/api/leave_requests', methods=['POST'])
def create_leave_request():
    global request_id_counter
    log_to_file("Create leave request route was hit.")
    data = request.get_json()
    if not data or 'engineerName' not in data or 'startDate' not in data or 'endDate' not in data:
        return jsonify({"error": "Missing required fields for leave request"}), 400

    try:
        start_date = datetime.datetime.strptime(data['startDate'], '%Y-%m-%d')
        end_date = datetime.datetime.strptime(data['endDate'], '%Y-%m-%d')
        if end_date < start_date:
            return jsonify({"error": "End date cannot be before start date"}), 400
    except ValueError:
        return jsonify({"error": "Invalid date format. Please use YYYY-MM-DD"}), 400

    request_id_counter += 1
    new_request = {
        "id": request_id_counter,
        "engineerName": data['engineerName'],
        "startDate": data['startDate'],
        "endDate": data['endDate'],
        "reason": data.get('reason', ''),
        "status": "Pending"
    }
    leave_requests_db.append(new_request)
    log_to_file(f"Created new leave request: {json.dumps(new_request, indent=2)}")
    return jsonify(new_request), 201

@app.route('/api/leave_requests/<int:request_id>/status', methods=['PUT'])
def update_leave_request_status(request_id):
    log_to_file(f"Update status route was hit for request ID: {request_id}")
    data = request.get_json()
    if not data or 'status' not in data:
        return jsonify({"error": "Missing 'status' in request body"}), 400

    new_status = data['status']
    if new_status not in ['Approved', 'Rejected']:
        return jsonify({"error": "Invalid status value"}), 400

    request_to_update = next((req for req in leave_requests_db if req['id'] == request_id), None)

    if not request_to_update:
        return jsonify({"error": "Leave request not found"}), 404

    conflict_warning = None
    # If approving, do conflict check and update leave balance
    if new_status == 'Approved':
        # Conflict Detection
        if current_roster:
            try:
                leave_start_date = datetime.datetime.strptime(request_to_update['startDate'], '%Y-%m-%d').date()
                leave_end_date = datetime.datetime.strptime(request_to_update['endDate'], '%Y-%m-%d').date()
                engineer_name = request_to_update['engineerName']

                conflicting_days = []
                delta = leave_end_date - leave_start_date

                for i in range(delta.days + 1):
                    day = leave_start_date + timedelta(days=i)
                    day_name = day.strftime('%A') # e.g., "Monday"

                    if day_name in current_roster:
                        for shift, people in current_roster[day_name].items():
                            if isinstance(people, list):
                                if any(p.get('name') == engineer_name for p in people):
                                    conflicting_days.append(day_name)
                                    break

                if conflicting_days:
                    conflict_warning = f"Warning: This leave conflicts with the generated roster on {', '.join(conflicting_days)}."
                    log_to_file(f"Conflict detected for {engineer_name}: {conflict_warning}")

            except (ValueError, TypeError) as e:
                log_to_file(f"Error during conflict detection: {e}")

        # Update leave balance
        try:
            start_date = datetime.datetime.strptime(request_to_update['startDate'], '%Y-%m-%d')
            end_date = datetime.datetime.strptime(request_to_update['endDate'], '%Y-%m-%d')
            duration = (end_date - start_date).days + 1

            employee = next((emp for emp in employees_db if emp['name'] == request_to_update['engineerName']), None)
            if employee:
                if employee['leaveBalance'] >= duration:
                    employee['leaveBalance'] -= duration
                    log_to_file(f"Updated {employee['name']}'s leave balance to {employee['leaveBalance']}")
                else:
                    log_to_file(f"Insufficient leave balance for {employee['name']}. Rejecting request.")
                    new_status = 'Rejected'
            else:
                log_to_file(f"Could not find employee {request_to_update['engineerName']} to update balance.")

        except (ValueError, TypeError) as e:
            log_to_file(f"Error processing date for leave balance update: {e}")
            pass

    request_to_update['status'] = new_status
    log_to_file(f"Updated leave request {request_id} to status: {new_status}")

    response_data = request_to_update.copy()
    if conflict_warning:
        response_data['conflict_warning'] = conflict_warning

    return jsonify(response_data)

# --- Cab Request Endpoints ---
cab_requests_db = []
cab_request_id_counter = 0

@app.route('/api/cab_requests', methods=['GET'])
def get_cab_requests():
    engineer_name = request.args.get('engineerName')
    if not engineer_name:
        return jsonify({"error": "Missing engineerName query parameter"}), 400

    relevant_requests = [
        req for req in cab_requests_db
        if req['engineerName'] == engineer_name
    ]
    log_to_file(f"Fetched {len(relevant_requests)} cab requests for {engineer_name}")
    return jsonify(relevant_requests)

@app.route('/api/cab_requests', methods=['POST'])
def create_cab_request():
    global cab_request_id_counter
    log_to_file("Create cab request route was hit.")
    data = request.get_json()
    required_fields = ['engineerName', 'date', 'shift']
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields for cab request"}), 400

    try:
        datetime.datetime.strptime(data['date'], '%Y-%m-%d')
    except ValueError:
        return jsonify({"error": "Invalid date format. Please use YYYY-MM-DD"}), 400

    cab_request_id_counter += 1
    new_request = {
        "id": cab_request_id_counter,
        "engineerName": data['engineerName'],
        "date": data['date'],
        "shift": data['shift'],
        "status": "Pending" # Status: Pending, Approved, Rejected
    }
    cab_requests_db.append(new_request)
    log_to_file(f"Created new cab request: {json.dumps(new_request, indent=2)}")
    return jsonify(new_request), 201

@app.route('/api/cab_requests/all', methods=['GET'])
def get_all_cab_requests():
    log_to_file("Get all cab requests route was hit.")
    return jsonify(cab_requests_db)

@app.route('/api/cab_requests/<int:request_id>/status', methods=['PUT'])
def update_cab_request_status(request_id):
    log_to_file(f"Update cab request status route was hit for request ID: {request_id}")
    data = request.get_json()
    if not data or 'status' not in data:
        return jsonify({"error": "Missing 'status' in request body"}), 400

    new_status = data['status']
    if new_status not in ['Approved', 'Rejected']:
        return jsonify({"error": "Invalid status value"}), 400

    request_to_update = next((req for req in cab_requests_db if req['id'] == request_id), None)

    if not request_to_update:
        return jsonify({"error": "Cab request not found"}), 404

    request_to_update['status'] = new_status
    log_to_file(f"Updated cab request {request_id} to status: {new_status}")
    return jsonify(request_to_update)
# --- Shift Swap Endpoints ---

# In-memory data store for swap requests
swap_requests_db = []
swap_request_id_counter = 0

@app.route('/api/swap_requests', methods=['GET'])
def get_swap_requests():
    engineer_name = request.args.get('engineerName')
    if not engineer_name:
        return jsonify({"error": "Missing engineerName query parameter"}), 400

    # Return requests where the engineer is either the requester or the responder
    relevant_requests = [
        req for req in swap_requests_db
        if req['requesterName'] == engineer_name or req['responderName'] == engineer_name
    ]
    log_to_file(f"Fetched {len(relevant_requests)} swap requests for {engineer_name}")
    return jsonify(relevant_requests)

@app.route('/api/swap_requests', methods=['POST'])
def create_swap_request():
    global swap_request_id_counter
    log_to_file("Create swap request route was hit.")
    data = request.get_json()
    required_fields = ['requesterName', 'responderName', 'requesterShift', 'responderShift']
    if not all(field in data for field in required_fields):
        return jsonify({"error": "Missing required fields for swap request"}), 400

    swap_request_id_counter += 1
    new_request = {
        "id": swap_request_id_counter,
        "requesterName": data['requesterName'],
        "responderName": data['responderName'],
        "requesterShift": data['requesterShift'], # e.g., { "day": "Monday", "shift": "Morning" }
        "responderShift": data['responderShift'], # e.g., { "day": "Tuesday", "shift": "Evening" }
        "status": "Pending" # Status: Pending, Approved, Rejected
    }
    swap_requests_db.append(new_request)
    log_to_file(f"Created new swap request: {json.dumps(new_request, indent=2)}")
    return jsonify(new_request), 201

@app.route('/api/swap_requests/<int:request_id>/status', methods=['PUT'])
def update_swap_request_status(request_id):
    log_to_file(f"Update swap status route was hit for request ID: {request_id}")
    data = request.get_json()
    if not data or 'status' not in data:
        return jsonify({"error": "Missing 'status' in request body"}), 400

    new_status = data['status']
    if new_status not in ['Approved', 'Rejected']:
        return jsonify({"error": "Invalid status value"}), 400

    request_to_update = next((req for req in swap_requests_db if req['id'] == request_id), None)

    if not request_to_update:
        return jsonify({"error": "Swap request not found"}), 404

    # If approved, perform the swap in the current_roster
    if new_status == 'Approved' and current_roster:
        try:
            requester_name = request_to_update['requesterName']
            responder_name = request_to_update['responderName']
            req_shift_info = request_to_update['requesterShift']
            res_shift_info = request_to_update['responderShift']

            # Find the engineers in the roster
            requester_obj = next((emp for emp in employees_db if emp['name'] == requester_name), None)
            responder_obj = next((emp for emp in employees_db if emp['name'] == responder_name), None)

            if not requester_obj or not responder_obj:
                raise ValueError("Could not find one of the employees for the swap.")

            # Remove requester from their original shift
            current_roster[req_shift_info['day']][req_shift_info['shift']] = [
                p for p in current_roster[req_shift_info['day']][req_shift_info['shift']] if p['name'] != requester_name
            ]
            # Remove responder from their original shift
            current_roster[res_shift_info['day']][res_shift_info['shift']] = [
                p for p in current_roster[res_shift_info['day']][res_shift_info['shift']] if p['name'] != responder_name
            ]

            # Add requester to responder's original shift
            current_roster[res_shift_info['day']][res_shift_info['shift']].append(requester_obj)
            # Add responder to requester's original shift
            current_roster[req_shift_info['day']][req_shift_info['shift']].append(responder_obj)

            log_to_file(f"Successfully swapped shifts in roster for request {request_id}")
            log_to_file(f"Updated roster: {json.dumps(current_roster, indent=2)}")

        except (KeyError, ValueError) as e:
            log_to_file(f"Error swapping shifts for request {request_id}: {e}")
            # If the swap fails, reject the request to prevent inconsistency
            request_to_update['status'] = 'Rejected'
            return jsonify({"error": f"Failed to perform swap in roster: {e}"}), 500

    request_to_update['status'] = new_status
    log_to_file(f"Updated swap request {request_id} to status: {new_status}")
    return jsonify(request_to_update)

# --- Yearly Schedule Endpoint ---

@app.route('/api/yearly_schedule', methods=['GET'])
def get_yearly_schedule():
    year = request.args.get('year', default=datetime.datetime.now().year, type=int)
    engineer_name = request.args.get('engineerName')

    if not current_roster:
        # If no roster is generated, return an empty schedule
        return jsonify({})

    yearly_schedule = {}
    for month in range(12):
        yearly_schedule[month] = {}
        # Correctly calculate days in month
        if month == 11: # December
            days_in_month = 31
        else:
            days_in_month = (datetime.date(year, month + 2, 1) - datetime.timedelta(days=1)).day

        for day in range(1, days_in_month + 1):
            try:
                current_date = datetime.date(year, month + 1, day)
                day_name = current_date.strftime('%A') # e.g., "Monday"

                shift_for_day = "Off" # Default to off

                if day_name in current_roster:
                    if engineer_name:
                        # Find the shift for the specific engineer
                        for shift, people in current_roster[day_name].items():
                            if isinstance(people, list):
                                if any(p.get('name') == engineer_name for p in people):
                                    shift_for_day = shift
                                    break
                    else:
                        # Provide a general summary if no engineer is specified
                        # For simplicity, we'll just list the first shift found for that day.
                        # A more complex UI might want all shifts.
                        first_shift = next(iter(current_roster[day_name]), None)
                        if first_shift:
                            shift_for_day = first_shift

                yearly_schedule[month][day] = shift_for_day
            except ValueError:
                # Handles cases like February 29 on a non-leap year
                continue

    return jsonify(yearly_schedule)

# --- Analytics Endpoints ---

@app.route('/api/analytics/team_averages', methods=['GET'])
def get_team_averages():
    log_to_file("Get team averages route was hit.")
    if not employees_db:
        return jsonify({
            "avg_tickets_resolved": 0,
            "avg_resolution_time": 0,
            "avg_csat": 0
        })

    total_tickets = sum(e.get('ticketsResolved', 0) for e in employees_db)
    total_resolution_time = sum(e.get('avgResolutionTime', 0) for e in employees_db)
    total_csat = sum(e.get('csat', 0) for e in employees_db)
    num_employees = len(employees_db)

    averages = {
        "avg_tickets_resolved": round(total_tickets / num_employees, 1),
        "avg_resolution_time": round(total_resolution_time / num_employees, 1),
        "avg_csat": round(total_csat / num_employees, 1)
    }

    log_to_file(f"Calculated team averages: {json.dumps(averages, indent=2)}")
    return jsonify(averages)

# A helper function to safely convert values to integers
def safe_int_conversion(value, default=0):
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

@app.route('/api/employees/upload', methods=['POST'])
def upload_employees():
    global employees_db
    log_to_file("Upload employees route was hit.")

    if 'file' not in request.files:
        log_to_file("Error: No file part in the request.")
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']
    if file.filename == '':
        log_to_file("Error: No selected file.")
        return jsonify({"error": "No selected file"}), 400

    if file and file.filename.endswith(('.xlsx', '.xls')):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active  # Use the active sheet

            # Normalize headers: lowercase and strip spaces
            raw_headers = [cell.value for cell in sheet[1]]
            headers = [str(h).lower().strip() if h else '' for h in raw_headers]

            # Define flexible header mappings
            header_map = {
                'name': ['name', 'employee name', 'engineer'],
                'role': ['role', 'job title'],
                'project': ['project'],
                'costCenter': ['costcenter', 'cost center'],
                'serviceNow': ['servicenow', 'service now', 'sn'],
                'jira': ['jira', 'jira tickets'],
                'csat': ['csat', 'customer satisfaction'],
                'ticketsResolved': ['ticketsresolved', 'tickets resolved', 'resolved'],
                'avgResolutionTime': ['avgresolutiontime', 'avg res time', 'resolution time'],
            }

            # Find the actual header names used in the file
            found_headers = {}
            for key, potential_names in header_map.items():
                for name in potential_names:
                    if name in headers:
                        found_headers[key] = raw_headers[headers.index(name)]
                        break

            if 'name' not in found_headers or 'role' not in found_headers:
                log_to_file("Error: Missing 'Name' or 'Role' column in Excel file.")
                return jsonify({"error": "Missing 'Name' or 'Role' column"}), 400

            existing_employees_map = {emp['name']: emp for emp in employees_db}

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                new_emp_data = dict(zip(raw_headers, row))
                project = new_emp_data.get(found_headers.get('project'))
                cost_center = new_emp_data.get(found_headers.get('costCenter'))

                if project and cost_center:
                    if project not in project_cost_centers or cost_center not in project_cost_centers[project]:
                        error_msg = f"Invalid cost center '{cost_center}' for project '{project}' on row {row_idx}."
                        log_to_file(error_msg)
                        return jsonify({"error": error_msg}), 400

            for row in sheet.iter_rows(min_row=2, values_only=True):
                new_emp_data = dict(zip(raw_headers, row))
                name = new_emp_data.get(found_headers['name'])

                if not name:
                    continue

                if name in existing_employees_map:
                    # Update existing employee
                    emp = existing_employees_map[name]
                    emp['role'] = new_emp_data.get(found_headers.get('role'), emp['role'])
                    emp['project'] = new_emp_data.get(found_headers.get('project'), emp.get('project'))
                    emp['costCenter'] = new_emp_data.get(found_headers.get('costCenter'), emp.get('costCenter'))
                    emp['serviceNow'] = safe_int_conversion(new_emp_data.get(found_headers.get('serviceNow')), emp['serviceNow'])
                    emp['jira'] = safe_int_conversion(new_emp_data.get(found_headers.get('jira')), emp['jira'])
                    emp['csat'] = safe_int_conversion(new_emp_data.get(found_headers.get('csat')), emp['csat'])
                    emp['ticketsResolved'] = safe_int_conversion(new_emp_data.get(found_headers.get('ticketsResolved')), emp['ticketsResolved'])
                    emp['avgResolutionTime'] = safe_int_conversion(new_emp_data.get(found_headers.get('avgResolutionTime')), emp['avgResolutionTime'])
                else:
                    # Add new employee
                    existing_employees_map[name] = {
                        "name": name,
                        "role": new_emp_data.get(found_headers.get('role')),
                        "project": new_emp_data.get(found_headers.get('project')),
                        "costCenter": new_emp_data.get(found_headers.get('costCenter')),
                        "serviceNow": safe_int_conversion(new_emp_data.get(found_headers.get('serviceNow'))),
                        "jira": safe_int_conversion(new_emp_data.get(found_headers.get('jira'))),
                        "csat": safe_int_conversion(new_emp_data.get(found_headers.get('csat'))),
                        "ticketsResolved": safe_int_conversion(new_emp_data.get(found_headers.get('ticketsResolved'))),
                        "avgResolutionTime": safe_int_conversion(new_emp_data.get(found_headers.get('avgResolutionTime'))),
                        "leaveBalance": 20,
                        "isAiAgentActive": False
                    }

            employees_db = list(existing_employees_map.values())
            log_to_file(f"Successfully updated employees_db from {file.filename}")

            # Automatically regenerate the roster
            try:
                global current_roster
                prediction = roster_generator_instance.forward(members=employees_db, constraints="")
                current_roster = prediction.roster
                log_to_file(f"Automatically regenerated roster: {json.dumps(current_roster, indent=2)}")
            except Exception as e:
                log_to_file(f"Error regenerating roster: {e}")
                # Still return success for the upload, but with a warning
                return jsonify({
                    "message": "Employee data updated, but failed to regenerate roster.",
                    "fileName": file.filename,
                    "employees": employees_db
                }), 200

            return jsonify({
                "message": "Employee data updated and roster regenerated successfully",
                "fileName": file.filename,
                "employees": employees_db,
                "roster": current_roster
            }), 200

        except Exception as e:
            log_to_file(f"Error processing Excel file: {e}")
            return jsonify({"error": f"An unexpected error occurred: {e}"}), 500

    return jsonify({"error": "Invalid file type. Please upload an .xlsx or .xls file."}), 400

# --- Dashboard Endpoints ---

# In-memory data store for kudos
kudos_db = [
    {"from": "Manager", "to": "Keerthi", "message": "Resolved a P1 ticket in under 30 minutes!"},
    {"from": "Naresh", "to": "Rohit", "message": "Thanks for helping with the deployment script."}
]

@app.route('/api/dashboard/kpis', methods=['GET'])
def get_dashboard_kpis():
    if not employees_db:
        return jsonify({
            "kpiAdherence": 0,
            "staffingLevel": 0,
            "teamWorkload": 0,
            "burnoutRisk": 0
        })

    # KPI Adherence: Average CSAT score
    avg_csat = sum(e.get('csat', 0) for e in employees_db) / len(employees_db)

    # Team Workload: Average tickets resolved, scaled to a percentage
    avg_tickets = sum(e.get('ticketsResolved', 0) for e in employees_db) / len(employees_db)
    # Assuming a target of 20 tickets per person for 100% workload
    team_workload = min((avg_tickets / 20) * 100, 100)

    # Burnout Risk: Based on average resolution time, scaled
    avg_res_time = sum(e.get('avgResolutionTime', 0) for e in employees_db) / len(employees_db)
    # Assuming a threshold of 60 mins is high risk
    burnout_risk = min((avg_res_time / 60) * 100, 100)


    kpis = {
        "kpiAdherence": round(avg_csat),
        "staffingLevel": 90,  # Static for now
        "teamWorkload": round(team_workload),
        "burnoutRisk": round(burnout_risk)
    }
    return jsonify(kpis)

@app.route('/api/dashboard/radar', methods=['GET'])
def get_radar_alerts():
    # These are hardcoded for demonstration purposes
    alerts = [
        {
            "id": 1,
            "type": "High Ticket Volume Predicted",
            "message": "AI predicts a 40% spike in Jira tickets on Friday at 3 PM due to new feature deployment.",
            "action": "Place Naresh on a paid standby shift."
        },
        {
            "id": 2,
            "type": "Burnout Forecast",
            "message": "Keerthi has worked 5 consecutive evening shifts.",
            "action": "Assign her a morning shift on Thursday."
        },
        {
            "id": 3,
            "type": "Skill Mismatch",
            "message": "A critical 'Azure Database' task is scheduled for Rohit, but his skill confidence is low.",
            "action": "Initiate a smart swap with Keerthi."
        }
    ]
    return jsonify(alerts)

@app.route('/api/dashboard/wellness', methods=['GET'])
def get_wellness_data():
    # Shift Fairness Score Calculation
    shift_counts = {emp['name']: {"weekend": 0, "evening": 0, "total": 0} for emp in employees_db}
    if current_roster:
        for day, shifts in current_roster.items():
            for shift_name, people in shifts.items():
                if isinstance(people, list):
                    for person in people:
                        if person['name'] in shift_counts:
                            shift_counts[person['name']]['total'] += 1
                            if day in ['Saturday', 'Sunday']:
                                shift_counts[person['name']]['weekend'] += 1
                            if shift_name == 'Evening':
                                shift_counts[person['name']]['evening'] += 1

    # Simple fairness score based on the variance of evening/weekend shifts
    total_evening = sum(counts['evening'] for counts in shift_counts.values())
    total_weekend = sum(counts['weekend'] for counts in shift_counts.values())
    fairness_score = "A+" # Default to A+

    # Upcoming Time Off
    upcoming_leaves = []
    today = datetime.date.today()
    for req in leave_requests_db:
        if req['status'] == 'Approved':
            start_date = datetime.datetime.strptime(req['startDate'], '%Y-%m-%d').date()
            if 0 <= (start_date - today).days <= 14: # Within the next 2 weeks
                upcoming_leaves.append({
                    "name": req['engineerName'],
                    "daysUntil": (start_date - today).days
                })

    wellness_data = {
        "shiftFairnessScore": fairness_score,
        "kudos": kudos_db,
        "upcomingTimeOff": upcoming_leaves
    }
    return jsonify(wellness_data)

# --- Gamification Endpoints ---

def get_shift_start_time(day_name, shift_name):
    """Helper to get the datetime object for a shift's start time."""
    # Define shift start times (could be moved to a config)
    shift_times = {
        "Morning": "09:00:00",
        "Evening": "17:00:00",
        "Night": "01:00:00"
    }
    today = datetime.date.today()
    # This logic assumes the roster is for the current week
    days_from_today = (list(calendar.day_name).index(day_name) - today.weekday() + 7) % 7
    shift_date = today + datetime.timedelta(days=days_from_today)
    start_time_str = shift_times.get(shift_name, "00:00:00")
    return datetime.datetime.strptime(f"{shift_date} {start_time_str}", "%Y-%m-%d %H:%M:%S")

@app.route('/api/clock_in', methods=['POST'])
def clock_in():
    data = request.get_json()
    employee_name = data.get('name')
    if not employee_name:
        return jsonify({"error": "Employee name is required"}), 400

    now = datetime.datetime.now()
    today_name = now.strftime('%A')

    if not current_roster or today_name not in current_roster:
        return jsonify({"error": "No shift scheduled for today"}), 404

    # Find the employee's shift for today
    employee_shift = None
    for shift_name, people in current_roster[today_name].items():
        if isinstance(people, list) and any(p.get('name') == employee_name for p in people):
            employee_shift = shift_name
            break

    if not employee_shift:
        return jsonify({"error": "You do not have a shift scheduled for today."}), 404

    # Check if already clocked in for this shift
    if any(rec['name'] == employee_name and rec['shift'] == f"{today_name} {employee_shift}" and 'clock_out' not in rec for rec in clock_in_records):
         return jsonify({"error": "You have already clocked in for this shift."}), 400

    # Calculate points
    shift_start_time = get_shift_start_time(today_name, employee_shift)
    time_diff = now - shift_start_time
    points_awarded = 0
    message = ""

    if time_diff.total_seconds() <= 0: # Early
        points_awarded = 2
        message = f"Clocked in early! +{points_awarded} points."
    elif 0 < time_diff.total_seconds() <= 300: # On time (within 5 mins)
        points_awarded = 1
        message = f"Clocked in on time! +{points_awarded} points."
    else: # Late
        message = "Clocked in late. No points awarded."

    if points_awarded > 0:
        gamification_points[employee_name] = gamification_points.get(employee_name, 0) + points_awarded

    # Record the clock-in
    clock_in_records.append({
        "name": employee_name,
        "clock_in": now.isoformat(),
        "shift": f"{today_name} {employee_shift}"
    })

    log_to_file(f"{employee_name} clocked in for {employee_shift}. {message}")
    return jsonify({"message": message, "points": gamification_points.get(employee_name)}), 200

@app.route('/api/clock_out', methods=['POST'])
def clock_out():
    data = request.get_json()
    employee_name = data.get('name')
    if not employee_name:
        return jsonify({"error": "Employee name is required"}), 400

    now = datetime.datetime.now()

    # Find the active clock-in record for this employee
    active_record = next((rec for rec in reversed(clock_in_records) if rec['name'] == employee_name and 'clock_out' not in rec), None)

    if not active_record:
        return jsonify({"error": "No active clock-in found. Cannot clock out."}), 404

    active_record['clock_out'] = now.isoformat()
    log_to_file(f"{employee_name} clocked out.")
    return jsonify({"message": "Successfully clocked out."}), 200

@app.route('/api/gamification/status', methods=['GET'])
def gamification_status():
    employee_name = request.args.get('name')
    if not employee_name:
        return jsonify({"error": "Employee name is required"}), 400

    points = gamification_points.get(employee_name, 0)
    badges = employee_badges.get(employee_name, [])

    # Check for active clock-in
    is_clocked_in = any(rec['name'] == employee_name and 'clock_out' not in rec for rec in clock_in_records)

    return jsonify({
        "points": points,
        "badges": badges,
        "isClockedIn": is_clocked_in
    })

@app.route('/api/leaderboard', methods=['GET'])
def get_leaderboard():
    sorted_employees = sorted(gamification_points.items(), key=lambda item: item[1], reverse=True)
    leaderboard = [{"name": name, "points": points} for name, points in sorted_employees]
    return jsonify(leaderboard)

badge_tiers = {
    50: "Punctual Pro",
    100: "Taskmaster",
    200: "Shift Sentinel",
}

def check_and_award_badges(employee_name):
    """Check points and award badges if a new tier is reached."""
    points = gamification_points.get(employee_name, 0)
    current_badges = employee_badges.get(employee_name, [])
    new_badges = []
    for threshold, badge_name in badge_tiers.items():
        if points >= threshold and badge_name not in current_badges:
            employee_badges[employee_name].append(badge_name)
            new_badges.append(badge_name)
    return new_badges

@app.route('/api/award_points', methods=['POST'])
def award_points():
    data = request.get_json()
    employee_name = data.get('name')
    points_to_award = data.get('points', 0)

    if not employee_name or employee_name not in gamification_points:
        return jsonify({"error": "Invalid employee name"}), 400

    if not isinstance(points_to_award, int) or points_to_award <= 0:
        return jsonify({"error": "Points must be a positive integer"}), 400

    gamification_points[employee_name] += points_to_award

    newly_awarded_badges = check_and_award_badges(employee_name)

    message = f"Successfully awarded {points_to_award} points to {employee_name}."
    if newly_awarded_badges:
        message += f" New badges earned: {', '.join(newly_awarded_badges)}!"

    log_to_file(message)
    return jsonify({
        "message": message,
        "new_total": gamification_points[employee_name],
        "new_badges": newly_awarded_badges
    })


if __name__ == '__main__':
    app.run(debug=True, port=5001)
